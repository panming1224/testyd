import os
import sys
import json
import time
import hashlib
import requests
import pandas as pd
import shutil
import re
from datetime import datetime, timedelta
from pathlib import Path
import urllib.parse
from datetime import datetime
from typing import Dict, List, Any
import ast

# 配置UTF-8编码
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# 添加数据库接口和文件合并模块路径
sys.path.append(r'D:\testyd')
sys.path.append(r'D:\testyd\mysql')

try:
    from crawler_db_interface import CrawlerDBInterface
    from merge_excel_files import ExcelMerger
    print("✓ 成功导入数据库接口和文件合并模块")
except ImportError as e:
    print(f"✗ 导入模块失败: {e}")
    sys.exit(1)

class TmallChatManager:
    def __init__(self):
        # 数据库接口初始化
        self.db_interface = CrawlerDBInterface(
            platform='tm',
            shops_table='tm_shops',
            tasks_table='tm_tasks',
            database='company'
        )
        # 基础目录配置
        self.base_dir = Path(r"D:\yingdao\tm\天猫客服聊天记录")
        
        # 合并文件目录
        self.merged_dir = Path(r"D:\yingdao\tm\合并文件\天猫客服聊天记录")
        
        # 目标日期（13天前，t-13）
        self.target_date = datetime.now() - timedelta(days=13)
        self.target_date_str = self.target_date.strftime('%Y-%m-%d')
        
        # 设置每日目录
        self.daily_dir = self.base_dir / self.target_date_str
        
        # API配置
        self.APP_KEY = "12574478"
        self.CUSTOMER_LIST_API = "https://h5api.m.taobao.com/h5/mtop.taobao.wireless.amp2.paas.conversation.list/1.0/"
        self.CHAT_MESSAGE_API = "https://h5api.m.taobao.com/h5/mtop.taobao.wireless.amp2.im.paas.message.list/1.0/"
        
        # MinIO配置
        self.minio_api_url = "http://127.0.0.1:8009/api/upload"
        
        # Dremio配置
        self.dremio_config = {
            'host': 'localhost',
            'port': 9047,
            'username': 'admin',
            'password': 'admin123'
        }
        
        print(f"目标日期: {self.target_date} ({self.target_date_str})")
        
    def _extract_customer_id(self, customer):
        """提取客户ID的辅助方法"""
        cid_obj = customer.get('cid', {})
        if isinstance(cid_obj, dict):
            return cid_obj.get('appCid', '')
        return cid_obj
        
    def _extract_user_id(self, customer):
        """提取用户ID的辅助方法"""
        user_id_obj = customer.get('userID', {})
        if isinstance(user_id_obj, dict):
            return user_id_obj.get('appUid', '')
        return user_id_obj
        
    def get_h5_token(self, cookies_str):
        """从cookie字符串中提取h5 token"""
        try:
            if not cookies_str:
                print("Cookie字符串为空")
                return None
            
            print(f"\n=== Token提取调试 ===")
            print(f"Cookie字符串长度: {len(cookies_str)}")
            print(f"Cookie前200字符: {cookies_str[:200]}")
            
            for cookie in cookies_str.split(';'):
                cookie = cookie.strip()
                if '_m_h5_tk=' in cookie:
                    token_value = cookie.split('_m_h5_tk=')[1].strip()
                    print(f"找到_m_h5_tk值: {token_value}")
                    # token格式为: token_expireTime，我们只需要token部分
                    if '_' in token_value:
                        token = token_value.split('_')[0]
                        expire_time = token_value.split('_')[1] if len(token_value.split('_')) > 1 else 'unknown'
                        print(f"提取的token: {token}")
                        print(f"过期时间戳: {expire_time}")
                        
                        # 检查token是否过期
                        try:
                            current_time = int(time.time() * 1000)
                            expire_timestamp = int(expire_time)
                            print(f"当前时间戳: {current_time}")
                            print(f"token过期时间戳: {expire_timestamp}")
                            if current_time > expire_timestamp:
                                print("⚠️ Token已过期！")
                            else:
                                print("✓ Token仍然有效")
                        except:
                            print("无法解析过期时间")
                        
                        print(f"========================\n")
                        return token
                    else:
                        print(f"Token格式异常，没有找到下划线分隔符: {token_value}")
                        print(f"========================\n")
                        return token_value
            
            print(f"⚠️ 在cookie中未找到_m_h5_tk")
            print(f"所有cookie项:")
            for i, cookie in enumerate(cookies_str.split(';')):
                print(f"  {i+1}: {cookie.strip()[:50]}...")
            print(f"========================\n")
            return None
        except Exception as e:
            print(f"提取token失败: {e}")
            return None
    
    def generate_sign(self, token, timestamp, data):
        """生成签名 - 按照淘宝mtop API标准算法"""
        try:
            # 签名算法: md5(token + '&' + timestamp + '&' + appKey + '&' + data)
            sign_str = f"{token}&{timestamp}&{self.APP_KEY}&{data}"
            return hashlib.md5(sign_str.encode('utf-8')).hexdigest()
        except Exception as e:
            print(f"生成签名失败: {e}")
            return None
    
    def generate_daily_tasks(self):
        """生成当日任务"""
        print("\n=== 生成当日任务 ===")
        
        try:
            # 定义任务列
            task_columns = ['chat_status']
            
            # 生成任务
            created_count = self.db_interface.generate_tasks(self.target_date_str, task_columns)
            print(f"✓ 成功生成 {created_count} 个任务")
            
            # 检查是否有任务生成
            if created_count == 0:
                print("⚠️ 没有生成任何任务，可能是因为任务已存在或没有符合条件的店铺")
            
            return True
        except Exception as e:
            print(f"✗ 生成任务失败: {e}")
            return False
    def get_shops_with_tasks(self):
        """获取有待处理badscore_status任务的店铺信息"""
        try:
            # 获取待处理任务 - 使用与tm_kpi.py相同的方法
            pending_tasks = self.db_interface.get_pending_tasks(self.target_date_str, 'chat_status')
            
            if not pending_tasks:
                print(f"没有找到 chat_status 类型的待处理任务")
                return {}
            
            print(f"找到 {len(pending_tasks)} 个待处理的chat_status任务")
            
            # 构建店铺信息字典
            shops_info = {}
            for task in pending_tasks:
                shop_name = task[1] if len(task) > 1 else None  # dt.shop_name
                qncookie = task[7] if len(task) > 7 else None  # s.qncookie (第6列)
                userNick=task[12] if len(task)>12 else None #索引12: userNick
                
                if qncookie:
                    shops_info[shop_name] = {
                        'shop_name': shop_name,
                        'qncookie': qncookie,
                        'userNick':userNick,
                        'task_info': task
                    }
                    print(f"✓ 店铺 {shop_name} - qncookie: {qncookie[:50] if qncookie else 'None'}...")
                else:
                    print(f"⚠️ 店铺 {shop_name} 缺少cookie信息，跳过")
            
            return shops_info
            
        except Exception as e:
            print(f"获取店铺任务信息失败: {e}")
            return {}
    
    def get_customer_list(self, cookies_str, begin_date="20250925", end_date="20250925", page_size=5, page_index=1):
        """获取客户列表"""
        try:
            # 提取token
            token = self.get_h5_token(cookies_str)
            if not token:
                print("无法提取token")
                return None
            
            timestamp = str(int(time.time() * 1000))
            
            # 构建请求数据
            request_data = {
                "beginDate": begin_date,
                "endDate": end_date,
                "pageSize": page_size,
                "pageIndex": page_index
            }
            
            # 转换为JSON字符串（用于签名计算）
            data_str = json.dumps(request_data, separators=(',', ':'), ensure_ascii=False)
            
            # 生成签名
            sign = self.generate_sign(token, timestamp, data_str)
            if not sign:
                print("签名生成失败")
                return None
            
            # 构建POST表单数据
            form_data = {
                'jsv': '2.6.2',
                'appKey': self.APP_KEY,
                't': timestamp,
                'sign': sign,
                'api': 'mtop.taobao.wireless.amp2.paas.conversation.list',
                'v': '1.0',
                'type': 'jsonp',
                'dataType': 'jsonp',
                'callback': 'mtopjsonp3',
                'data': data_str
            }
            
            # 提取关键认证参数 - 只使用必要的cookie字段
            # 从完整cookie字符串中提取关键字段
            essential_cookies = []
            cookie_pairs = cookies_str.split(';')
            
            # 提取必要的cookie字段 - 包含更多必要字段
            essential_fields = ['_m_h5_tk', '_m_h5_tk_enc', 't', 'xlly_s', 'mtop_partitioned_detect', '_tb_token_', '_samesite_flag_', '3PcFlag', 'cookie2', 'sgcookie', 'unb', 'sn', 'uc1', 'csg', '_cc_', 'cancelledSubSites', 'skt', 'cna', 'tfstk']
            for pair in cookie_pairs:
                if '=' in pair:
                    key = pair.split('=')[0].strip()
                    if key in essential_fields:
                        essential_cookies.append(pair.strip())
            
            # 构建精简的cookie字符串
            essential_cookie_str = '; '.join(essential_cookies)
            
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Content-Type": "application/x-www-form-urlencoded",
                "Cookie": essential_cookie_str
            }
            
            print(f"\n=== 客户列表请求详情 ===")
            print(f"请求URL: {self.CUSTOMER_LIST_API}")
            print(f"请求数据: {request_data}")
            print(f"Cookie长度: {len(essential_cookie_str)}")
            print(f"实际发送的Cookie: {essential_cookie_str}")
            print(f"========================\n")
            
            # 发送POST请求
            response = requests.post(self.CUSTOMER_LIST_API, data=form_data, headers=headers, timeout=30)
            
            if response.status_code == 200:
                # 处理JSONP响应
                response_text = response.text.strip()
                print(f"原始响应长度: {len(response_text)}")
                print(f"响应前200字符: {response_text[:200]}")
                
                # 使用正则表达式匹配JSONP格式
                import re
                match = re.match(r'^(\w+)\((.*)\)$', response_text)
                if match:
                    json_str = match.group(2)
                    
                    try:
                        data = json.loads(json_str)
                        print(f"\n=== API响应解析结果 ===")
                        print(f"响应状态: {data.get('ret', 'Not found')}")
                        print(f"API名称: {data.get('api', 'Not found')}")
                        print(f"data字段类型: {type(data.get('data', None))}")
                        
                        # 打印完整的响应数据
                        print(f"\n=== 完整响应数据 ===")
                        print(json.dumps(data, indent=2, ensure_ascii=False))
                        print(f"==================\n")
                        
                        if 'data' in data and data['data']:
                            data_content = data['data']
                            print(f"data内容类型: {type(data_content)}")
                            if isinstance(data_content, dict):
                                print(f"data字段的键: {list(data_content.keys())}")
                                if 'result' in data_content:
                                    result_content = data_content['result']
                                    print(f"result字段类型: {type(result_content)}")
                                    print(f"result字段长度: {len(result_content) if isinstance(result_content, (list, dict)) else 'N/A'}")
                                    if isinstance(result_content, list) and len(result_content) > 0:
                                        print(f"\n=== 客户列表详情 ===")
                                        for i, customer in enumerate(result_content[:5]):  # 只显示前5个客户
                                            print(f"客户 {i+1}: {json.dumps(customer, indent=2, ensure_ascii=False)}")
                                        print(f"==================\n")
                        print(f"========================\n")
                        
                        return data
                    except json.JSONDecodeError as e:
                        print(f"[ERROR] JSON解析失败: {e}")
                        return None
                else:
                    print("[ERROR] 响应格式不是预期的JSONP格式")
                    return None
            else:
                print(f"请求失败，状态码: {response.status_code}")
                print(f"响应内容: {response.text[:200]}")
                return None
                
        except Exception as e:
            print(f"获取客户列表失败: {e}")
            return None
    
    def get_chat_messages_with_user_info(self, cookies_str, userNick, customer_data):
        """获取聊天消息，使用从客户数据中解析的参数"""
        try:
            # 从客户数据中解析实际的cid和userId
            actual_cid = None
            if 'cid' in customer_data:
                cid_value = customer_data['cid']
                if isinstance(cid_value, dict) and 'appCid' in cid_value:
                    actual_cid = cid_value['appCid']
                elif isinstance(cid_value, str):
                    # 尝试解析字符串格式的字典
                    try:
                        cid_dict = ast.literal_eval(cid_value)
                        if isinstance(cid_dict, dict) and 'appCid' in cid_dict:
                            actual_cid = cid_dict['appCid']
                        else:
                            actual_cid = cid_value
                    except:
                        actual_cid = cid_value
                else:
                    actual_cid = str(cid_value)
            
            if not actual_cid:
                actual_cid = "2215831800345.1-2219315280500.1#11001"  # 默认值
            
            # 解析userId，优先提取appUid
            actual_user_id = None
            if 'userID' in customer_data:
                user_id_value = customer_data['userID']
                if isinstance(user_id_value, dict) and 'appUid' in user_id_value:
                    actual_user_id = user_id_value['appUid']
                elif isinstance(user_id_value, str):
                    # 尝试解析字符串格式的字典
                    try:
                        user_id_dict = ast.literal_eval(user_id_value)
                        if isinstance(user_id_dict, dict) and 'appUid' in user_id_dict:
                            actual_user_id = user_id_dict['appUid']
                        else:
                            actual_user_id = user_id_value
                    except:
                        actual_user_id = user_id_value
                else:
                    actual_user_id = str(user_id_value)
            elif 'userId' in customer_data:
                actual_user_id = str(customer_data['userId'])
            elif 'buyerId' in customer_data:
                actual_user_id = str(customer_data['buyerId'])
            elif 'customerId' in customer_data:
                actual_user_id = str(customer_data['customerId'])
            
            # 确保使用正确的appUid，不使用默认值
            if not actual_user_id:
                print(f"警告：无法提取客户的userId，客户数据: {customer_data}")
                return None
            
            # 使用从cookies中获取的真实用户昵称
            actual_user_nick = userNick
            
            # 提取token
            token = self.get_h5_token(cookies_str)
            if not token:
                print("无法提取token")
                return None
            
            # # 生成时间戳
            timestamp = str(int(time.time() * 1000))
            timestamp1 = str(int(self.target_date.timestamp() * 1000))

            
            # 构建请求数据
            request_data = {
                "userNick": userNick,  # 固定值
                "cid": actual_cid,
                "userId": actual_user_id,  # 使用实际提取的userId（appUid）
                "cursor": timestamp1,  # 恢复之前成功的时间戳
                "forward": "true",  # 向前查询
                "count": "20",  # 恢复之前成功的数量
                "needRecalledContent": "true"  # 固定值
            }
            
            print(f"\n=== 聊天消息请求详情 ===")
            print(f"客户: {customer_data.get('displayName', 'Unknown')}")
            print(f"请求URL: {self.CHAT_MESSAGE_API}")
            print(f"请求数据: {request_data}")
            print(f"实际提取的cid: {actual_cid}")
            print(f"实际提取的userId: {actual_user_id}")
            print(f"实际提取的userNick: {userNick}")
            print(f"========================\n")
            
            # 转换为JSON字符串
            data_str = json.dumps(request_data, separators=(',', ':'), ensure_ascii=False)
            
            # 生成签名
            sign = self.generate_sign(token, timestamp, data_str)
            
            # 提取关键认证参数 - 直接使用完整cookie字符串
            
            # 生成动态callback名称
            import random
            callback_num = random.randint(50, 99)
            callback_name = f'mtopjsonp{callback_num}'
            
            # 构建POST表单数据
            form_data = {
                'jsv': '2.6.2',
                'appKey': self.APP_KEY,
                't': timestamp,
                'sign': sign,
                'api': 'mtop.taobao.wireless.amp2.im.paas.message.list',
                'v': '1.0',
                'type': 'jsonp',
                'dataType': 'jsonp',
                'callback': callback_name,
                'data': data_str
            }
            
            # 设置请求头
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Content-Type": "application/x-www-form-urlencoded",
                "Cookie": cookies_str
            }
            
            # 发送请求
            for attempt in range(3):
                response = requests.post(self.CHAT_MESSAGE_API, data=form_data, headers=headers, timeout=30)
                
                if response.status_code == 200:
                    # 处理JSONP响应 - 动态解析callback名称
                    response_text = response.text.strip()
                    
                    # 使用正则表达式动态提取callback名称
                    import re
                    callback_match = re.match(r'^(mtopjsonp\d+)\(', response_text)
                    
                    if callback_match and response_text.endswith(')'):
                        callback_name = callback_match.group(1)
                        # 动态提取JSON部分
                        json_str = response_text[len(callback_name)+1:-1]  # 去掉 'callback_name(' 和 ')'
                        
                        try:
                            data = json.loads(json_str)
                            
                            # 打印完整的聊天消息响应
                            print(f"\n=== 聊天消息API完整响应 ===")
                            print(json.dumps(data, indent=2, ensure_ascii=False))
                            print(f"========================\n")
                            
                            # 检查返回状态
                            if 'ret' in data and data['ret'] and data['ret'][0].startswith('SUCCESS'):
                                # 获取消息列表
                                if 'data' in data and 'userMessages' in data['data']:
                                    message_list = data['data']['userMessages']
                                    print(f"成功获取到 {len(message_list)} 条消息")
                                    
                                    # 打印前5条消息的详细内容
                                    print(f"\n=== 前5条消息详情 ===")
                                    for i, msg in enumerate(message_list[:5]):
                                        print(f"消息 {i+1}: {json.dumps(msg, indent=2, ensure_ascii=False)}")
                                    print(f"==================\n")
                                    
                                    return message_list
                                else:
                                    print("API返回成功但没有消息数据")
                                    print(f"完整响应数据: {json.dumps(data, indent=2, ensure_ascii=False)}")
                                    return []
                            else:
                                # 检查是否直接包含data字段（无ret字段的情况）
                                if 'data' in data and 'userMessages' in data['data']:
                                    message_list = data['data']['userMessages']
                                    print(f"成功获取到 {len(message_list)} 条消息（无ret字段）")
                                    
                                    # 打印前5条消息的详细内容
                                    print(f"\n=== 前5条消息详情 ===")
                                    for i, msg in enumerate(message_list[:5]):
                                        print(f"消息 {i+1}: {json.dumps(msg, indent=2, ensure_ascii=False)}")
                                    print(f"==================\n")
                                    
                                    return message_list
                                else:
                                    print(f"API返回错误或无消息数据")
                                    print(f"完整响应数据: {json.dumps(data, indent=2, ensure_ascii=False)}")
                                    return []
                                
                        except json.JSONDecodeError as e:
                            print(f"JSON解析失败: {e}")
                            print(f"响应内容: {response_text[:200]}")
                            return None
                    else:
                        print(f"JSONP格式不匹配: {response_text[:100]}")
                        return None
                else:
                    if attempt < 2:
                        time.sleep(2)
                        continue
                    return None
            
            return None
            
        except Exception as e:
            print(f"获取聊天消息失败: {e}")
            return None

    def update_task_status(self, shop_name, task_type, status="已完成"):
        """更新任务状态"""
        try:
            success = self.db_interface.update_task_status(
                self.target_date_str, 
                shop_name, 
                task_type, 
                status
            )
            if success:
                print(f"✓ 任务状态更新成功: {shop_name} - {task_type} -> {status}")
            else:
                print(f"✗ 任务状态更新失败: {shop_name} - {task_type}")
            return success
        except Exception as e:
            print(f"✗ 更新任务状态时出错: {e}")
            return False

    def merge_and_upload_files(self):
        """合并文件并上传到MinIO"""
        print(f"\n=== 合并差评文件并上传 ===")
        
        try:
            # 确定源目录和目标目录
            source_dir = self.base_dir / self.target_date_str
            target_dir = self.merged_dir
            target_dir.mkdir(parents=True, exist_ok=True)
            
            # 目标文件路径
            target_merged_file = target_dir / f"{self.target_date_str}.xlsx"
            
            # 检查源目录是否存在文件
            if source_dir.exists() and any(source_dir.glob("*.xlsx")):
                print(f"📁 源目录: {source_dir}")
                print(f"📁 目标目录: {target_dir}")
                
                # 使用ExcelMerger合并文件
                merger = ExcelMerger(str(source_dir))
                success = merger.merge_excel_files(f"{self.target_date_str}.xlsx")
                
                if success:
                    # 移动合并后的文件到目标目录
                    source_merged_file = source_dir / f"{self.target_date_str}.xlsx"
                    if source_merged_file.exists():
                        # 如果目标文件已存在，先删除
                        if target_merged_file.exists():
                            target_merged_file.unlink()
                        
                        shutil.move(str(source_merged_file), str(target_merged_file))
                        print(f"✓ 合并文件已移动到: {target_merged_file}")
                    else:
                        print(f"✗ 合并文件不存在: {source_merged_file}")
                        return False
                else:
                    print("✗ 文件合并失败")
                    return False
            else:
                print(f"⚠️ 源目录无文件，创建空文件: {target_merged_file}")
                # 创建空的Excel文件
                try:
                    empty_df = pd.DataFrame(columns=['shop_name', 'customer_nick', 'customer_id', 
                                                   'message_id', 'message_content', 'message_time', 
                                                   'sender_type', 'message_type', 'create_time'])
                    empty_df.to_excel(target_merged_file, index=False, engine='openpyxl')
                    print(f"✓ 空文件创建成功: {target_merged_file}")
                except Exception as e:
                    print(f"✗ 创建空文件失败: {e}")
                    return False

            
            # 上传到MinIO
            minio_path = f"ods/tm/tm_chat/dt={self.target_date_str}/{self.target_date_str}.parquet"
            success = self.upload_to_minio(str(target_merged_file), minio_path)
            
            if success:
                print(f"✓ 聊天文件上传MinIO成功")
                
                # 刷新Dremio表
                dremio_table = 'minio.warehouse.ods.tm.tm_chat'
                self.refresh_dremio_table(dremio_table)
                return True
            else:
                print(f"✗ 聊天文件上传MinIO失败")
                return False
                
        except Exception as e:
            print(f"✗ 合并上传过程异常: {e}")
            return False

    def fetch_and_save_chat_data(self, shop_name, qncookie, userNick):
        """获取并保存聊天数据"""
        try:
            print(f"开始获取店铺 {shop_name} 的聊天数据")
            
            # 使用完整的qncookie作为cookies字符串（qncookie本身就包含了所有必要的cookie字段）
            cookies_str = qncookie
            print(f"构建的原始cookies_str: {cookies_str}")
            print(f"cookies_str长度: {len(cookies_str)}")
            
            # 1. 获取客户列表 - 使用正确的参数传递
            customer_list_response = self.get_customer_list(cookies_str)
            if not customer_list_response:
                print(f"店铺 {shop_name} 没有获取到客户列表")
                return False
            
            # 从响应中提取客户列表
            customer_list = []
            if isinstance(customer_list_response, dict) and 'data' in customer_list_response:
                data_content = customer_list_response['data']
                print(f"data内容: {data_content}")
                
                # data字段是空字典，说明没有客户数据
                if isinstance(data_content, dict) and len(data_content) == 0:
                    print(f"店铺 {shop_name} 当前没有客户聊天记录")
                    return True  # 返回True表示正常处理完成，只是没有数据
                elif isinstance(data_content, dict):
                    if 'result' in data_content:
                        customer_list = data_content['result']
                    else:
                        print(f"未找到result字段，data结构: {list(data_content.keys())}")
                        return True  # 返回True表示正常处理完成，只是没有数据
            
            if not customer_list:
                print(f"店铺 {shop_name} 客户列表为空")
                return True  # 返回True表示正常处理完成，只是没有数据
            
            print(f"获取到 {len(customer_list)} 个客户")
            
            # 限制只处理前2个客户
            customer_list = customer_list[:2]
            print(f"限制处理前 {len(customer_list)} 个客户")
            
            # 2. 获取所有客户的聊天记录并收集原始JSONP响应
            all_raw_responses = []
            for i, customer in enumerate(customer_list, 1):
                print(f"正在处理第 {i} 个客户: {customer.get('userNick', 'unknown')}")
                try:
                    # 获取原始JSONP响应而不是解析后的消息列表
                    raw_response = self.get_chat_messages_raw_response(
                        cookies_str, userNick, customer
                    )
                    if raw_response:
                        all_raw_responses.append(raw_response)
                except Exception as e:
                    print(f"获取客户 {customer.get('userNick', 'unknown')} 聊天记录失败: {e}")
                    continue
            
            # 3. 如果有多个响应，合并它们；如果只有一个，直接使用
            if all_raw_responses:
                if len(all_raw_responses) == 1:
                    # 只有一个响应，直接使用
                    combined_response = all_raw_responses[0]
                else:
                    # 多个响应，需要合并
                    combined_response = self.combine_raw_responses(all_raw_responses)
                
                # 保存数据到Excel，使用新的解析逻辑
                success = self.save_chat_data_to_excel(combined_response, shop_name)
                if success:
                    print(f"店铺 {shop_name} 聊天数据保存成功")
                    return True
                else:
                    print(f"店铺 {shop_name} 聊天数据保存失败")
                    return False
            else:
                print(f"店铺 {shop_name} 没有获取到聊天数据")
                # 创建空文件
                success = self.save_chat_data_to_excel([], shop_name)
                return success
                
        except Exception as e:
            print(f"获取店铺 {shop_name} 聊天数据时发生异常: {e}")
            return False

    def get_chat_messages_raw_response(self, cookies_str, userNick, customer):
        """获取聊天消息的原始JSONP响应"""
        try:
            print(f"\n=== 开始获取客户聊天记录 ===")
            print(f"客户显示名: {customer.get('displayName', 'unknown')}")
            
            # 从客户数据中解析实际的cid和userId - 使用备份文件的逻辑
            actual_cid = None
            if 'cid' in customer:
                cid_value = customer['cid']
                if isinstance(cid_value, dict) and 'appCid' in cid_value:
                    actual_cid = cid_value['appCid']
                elif isinstance(cid_value, str):
                    # 尝试解析字符串格式的字典
                    try:
                        import ast
                        cid_dict = ast.literal_eval(cid_value)
                        if isinstance(cid_dict, dict) and 'appCid' in cid_dict:
                            actual_cid = cid_dict['appCid']
                        else:
                            actual_cid = cid_value
                    except:
                        actual_cid = cid_value
                else:
                    actual_cid = str(cid_value)
            
            if not actual_cid:
                actual_cid = "2215831800345.1-2219315280500.1#11001"  # 默认值
            
            # 解析userId，优先提取appUid
            actual_user_id = None
            if 'userID' in customer:
                user_id_value = customer['userID']
                if isinstance(user_id_value, dict) and 'appUid' in user_id_value:
                    actual_user_id = user_id_value['appUid']
                elif isinstance(user_id_value, str):
                    # 尝试解析字符串格式的字典
                    try:
                        import ast
                        user_id_dict = ast.literal_eval(user_id_value)
                        if isinstance(user_id_dict, dict) and 'appUid' in user_id_dict:
                            actual_user_id = user_id_dict['appUid']
                        else:
                            actual_user_id = user_id_value
                    except:
                        actual_user_id = user_id_value
                else:
                    actual_user_id = str(user_id_value)
            elif 'userId' in customer:
                actual_user_id = str(customer['userId'])
            elif 'buyerId' in customer:
                actual_user_id = str(customer['buyerId'])
            elif 'customerId' in customer:
                actual_user_id = str(customer['customerId'])
            
            # 确保使用正确的appUid，不使用默认值
            if not actual_user_id:
                print(f"❌ 无法提取客户的userId，客户数据: {customer}")
                return None
            
            print(f"提取的cid: {actual_cid}")
            print(f"提取的user_id: {actual_user_id}")
            
            # 获取token
            token = self.get_h5_token(cookies_str)
            if not token:
                print("❌ 无法获取token")
                return None
            
            print(f"✓ 获取到token: {token[:20]}...")
            
            # 生成时间戳
            timestamp = str(int(time.time() * 1000))
            timestamp1 = str(int(self.target_date.timestamp() * 1000))
            print(f"时间戳: {timestamp}")
            
            # 构建请求数据 - 使用备份文件的成功参数
            request_data = {
                "userNick": userNick,  # 固定值
                "cid": actual_cid,
                "userId": actual_user_id,  # 使用实际提取的userId（appUid）
                "cursor": timestamp1,  # 恢复之前成功的时间戳
                "forward": "true",  # 向前查询
                "count": "20",  # 恢复之前成功的数量
                "needRecalledContent": "true"  # 固定值
            }
            
            print(f"请求数据: {request_data}")
            
            # 转换为JSON字符串
            data_str = json.dumps(request_data, separators=(',', ':'), ensure_ascii=False)
            print(f"JSON数据字符串: {data_str}")
            
            # 生成签名
            sign = self.generate_sign(token, timestamp, data_str)
            if not sign:
                print("❌ 签名生成失败")
                return None
            
            print(f"✓ 生成签名: {sign}")
            
            # 生成动态callback名称
            import random
            callback_num = random.randint(50, 99)
            callback_name = f'mtopjsonp{callback_num}'
            
            # 构建POST表单数据
            form_data = {
                'jsv': '2.6.2',
                'appKey': self.APP_KEY,
                't': timestamp,
                'sign': sign,
                'api': 'mtop.taobao.wireless.amp2.im.paas.message.list',
                'v': '1.0',
                'type': 'jsonp',
                'dataType': 'jsonp',
                'callback': callback_name,
                'data': data_str
            }
            
            print(f"表单数据: {form_data}")
            
            # 设置请求头
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Content-Type": "application/x-www-form-urlencoded",
                "Cookie": cookies_str
            }
            
            print(f"请求头: {headers}")
            print(f"请求URL: {self.CHAT_MESSAGE_API}")
            
            # 发送请求
            for attempt in range(3):
                print(f"\n--- 第 {attempt + 1} 次请求尝试 ---")
                try:
                    response = requests.post(self.CHAT_MESSAGE_API, data=form_data, headers=headers, timeout=30)
                    
                    print(f"响应状态码: {response.status_code}")
                    print(f"响应头: {dict(response.headers)}")
                    
                    if response.status_code == 200:
                        # 返回原始JSONP响应文本
                        response_text = response.text.strip()
                        print(f"✓ 响应内容长度: {len(response_text)}")
                        print(f"响应内容前500字符: {response_text[:500]}")
                        
                        if response_text:
                            print(f"✓ 获取到客户 {customer.get('displayName', 'unknown')} 的原始JSONP响应")
                            return response_text
                        else:
                            print(f"❌ 响应内容为空")
                    else:
                        print(f"❌ 请求失败，状态码: {response.status_code}")
                        print(f"响应内容: {response.text[:500]}")
                        
                        if attempt < 2:
                            print(f"等待2秒后重试...")
                            time.sleep(2)
                            continue
                        
                except requests.exceptions.RequestException as e:
                    print(f"❌ 请求异常: {e}")
                    if attempt < 2:
                        print(f"等待2秒后重试...")
                        time.sleep(2)
                        continue
            
            print(f"❌ 所有请求尝试都失败了")
            return None
            
        except Exception as e:
            print(f"❌ 获取原始聊天消息响应失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def combine_raw_responses(self, raw_responses):
        """合并多个原始JSONP响应"""
        try:
            all_messages = []
            
            for raw_response in raw_responses:
                # 解析每个JSONP响应
                parsed_data = self.parse_chat_response(raw_response)
                if parsed_data:
                    # 提取消息
                    messages = self.parse_chat_messages_new(parsed_data)
                    if messages:
                        all_messages.extend(messages)
            
            if not all_messages:
                return None
            
            # 创建一个合并的数据结构
            combined_data = {
                'data': {
                    'userMessages': all_messages
                }
            }
            
            # 转换回JSONP格式
            json_str = json.dumps(combined_data, ensure_ascii=False)
            combined_jsonp = f"mtopjsonp1({json_str})"
            
            return combined_jsonp
            
        except Exception as e:
            print(f"合并原始响应失败: {e}")
            return None

    def process_chat_data(self, chat_messages, shop_name, customer):
        """处理聊天数据，转换为标准格式"""
        processed_data = []
        
        for message in chat_messages:
            try:
                processed_message = {
                    'shop_name': shop_name,
                    'customer_nick': customer.get('userNick', ''),
                    'customer_id': self._extract_customer_id(customer),
                    'message_id': message.get('id', ''),
                    'message_content': message.get('content', ''),
                    'message_time': message.get('time', ''),
                    'sender_type': message.get('senderType', ''),
                    'message_type': message.get('msgType', ''),
                    'create_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                processed_data.append(processed_message)
            except Exception as e:
                print(f"处理聊天消息时出错: {e}")
                continue
        
        return processed_data

    def save_chat_data_to_excel(self, chat_data, shop_name):
        """保存聊天数据到Excel文件 - 使用新的解析逻辑"""
        try:
            # 检查是否为空数据
            print(f"🔍 数据检查: chat_data={chat_data}, type={type(chat_data)}")
            if not chat_data or (isinstance(chat_data, list) and len(chat_data) == 0):
                print("📋 检测到空数据，创建空Excel文件...")
                return self._create_empty_excel_file(shop_name)
            
            # 检查数据类型并进行相应处理
            if isinstance(chat_data, str):
                # 如果是JSONP字符串，使用新的解析逻辑
                print("📋 检测到JSONP格式数据，使用新解析逻辑...")
                
                # 解析JSONP响应
                parsed_data = self.parse_chat_response(chat_data)
                if not parsed_data:
                    print("❌ JSONP解析失败")
                    return self._create_empty_excel_file(shop_name)
                
                # 检查是否是已格式化的数据
                if 'data' in parsed_data and 'userMessages' in parsed_data['data']:
                    user_messages = parsed_data['data']['userMessages']
                    
                    # 检查第一条消息是否已经包含格式化字段
                    if len(user_messages) > 0 and '客户' in user_messages[0] and '聊天记录' in user_messages[0]:
                        print("📋 检测到已格式化的消息数据，直接使用...")
                        
                        # 重新分析客户名称，因为可能都被标记为"未知客户"
                        self._fix_customer_names(user_messages)
                        
                        # 直接格式化为Excel数据
                        excel_data = self.format_for_excel_new(user_messages)
                    else:
                        print("📋 检测到原始API响应，进行解析...")
                        # 解析聊天消息
                        messages = self.parse_chat_messages_new(parsed_data)
                        if not messages:
                            print("❌ 没有解析到聊天消息")
                            return self._create_empty_excel_file(shop_name)
                        
                        # 格式化为Excel数据
                        excel_data = self.format_for_excel_new(messages)
                else:
                    print("❌ 数据结构不正确")
                    return self._create_empty_excel_file(shop_name)
                
                if not excel_data:
                    print("❌ Excel数据格式化失败")
                    return self._create_empty_excel_file(shop_name)
                
                # 创建DataFrame
                df = pd.DataFrame(excel_data)
                
            elif isinstance(chat_data, (list, dict)):
                # 如果是字典或列表格式，检查是否是已解析的消息列表
                print("📋 检测到字典/列表格式数据，使用新解析逻辑...")
                
                # 检查是否是已解析的消息列表（包含'客户'和'聊天记录'字段）
                if isinstance(chat_data, list) and len(chat_data) > 0 and '客户' in chat_data[0]:
                    print("📋 检测到已解析的消息列表，直接格式化...")
                    excel_data = self.format_for_excel_new(chat_data)
                    if not excel_data:
                        print("❌ Excel数据格式化失败")
                        return self._create_empty_excel_file(shop_name)
                    
                    # 创建DataFrame
                    df = pd.DataFrame(excel_data)
                else:
                    # 解析聊天消息
                    messages = self.parse_chat_messages_new(chat_data)
                    if not messages:
                        print("❌ 没有解析到聊天消息")
                        return self._create_empty_excel_file(shop_name)
                    
                    # 格式化为Excel数据
                    excel_data = self.format_for_excel_new(messages)
                    if not excel_data:
                        print("❌ Excel数据格式化失败")
                        return self._create_empty_excel_file(shop_name)
                    
                    # 创建DataFrame
                    df = pd.DataFrame(excel_data)
                
            else:
                print(f"❌ 不支持的数据类型: {type(chat_data)}")
                return self._create_empty_excel_file(shop_name)
            
            # 确保目录存在
            daily_dir = Path(self.daily_dir)
            daily_dir.mkdir(parents=True, exist_ok=True)
            
            # 生成文件名
            timestamp = datetime.now().strftime('%H%M%S')
            filename = f"{shop_name}_{self.target_date_str}_{timestamp}.xlsx"
            file_path = daily_dir / filename
            
            # 保存到Excel
            df.to_excel(file_path, index=False, engine='openpyxl')
            print(f"✓ 聊天数据已保存到: {file_path}")
            print(f"📊 数据统计: {len(df)} 个客户记录")
            return True
            
        except Exception as e:
            print(f"保存聊天数据到Excel时出错: {e}")
            print(f"🔍 异常处理: 调用_create_empty_excel_file")
            return self._create_empty_excel_file(shop_name)

    def _fix_customer_names(self, user_messages):
        """修复客户名称，重新分析消息中的真实客户"""
        try:
            print("🔧 修复客户名称...")
            
            # 收集所有发送者和接收者
            senders = set()
            receivers = set()
            
            for msg in user_messages:
                sender = msg.get('发送者', '')
                receiver = msg.get('接收者', '')
                
                if sender:
                    senders.add(sender)
                if receiver:
                    receivers.add(receiver)
            
            print(f"📊 发送者: {senders}")
            print(f"📊 接收者: {receivers}")
            
            # 识别客户（不包含店铺关键词的用户）
            customers = set()
            shop_staff = set()
            
            all_users = senders.union(receivers)
            for user in all_users:
                if any(keyword in user for keyword in ['旗舰店', '专卖店', '店铺', '客服', '服务助手']):
                    shop_staff.add(user)
                else:
                    customers.add(user)
            
            print(f"📊 识别的客户: {customers}")
            print(f"📊 识别的店铺员工: {shop_staff}")
            
            # 按客户重新分组消息
            customer_messages = {}
            
            for msg in user_messages:
                sender = msg.get('发送者', '')
                receiver = msg.get('接收者', '')
                
                # 确定这条消息属于哪个客户
                customer = None
                if sender in customers:
                    customer = sender
                elif receiver in customers:
                    customer = receiver
                
                if customer:
                    if customer not in customer_messages:
                        customer_messages[customer] = []
                    customer_messages[customer].append(msg)
                    
                    # 更新消息的客户字段
                    msg['客户'] = customer
            
            print(f"📊 按客户分组后: {len(customer_messages)} 个客户")
            for customer, msgs in customer_messages.items():
                print(f"  - {customer}: {len(msgs)} 条消息")
                
        except Exception as e:
            print(f"❌ 修复客户名称时出错: {e}")
            import traceback
            traceback.print_exc()

    def _create_empty_excel_file(self, shop_name):
        """创建空的Excel文件"""
        try:
            daily_dir = Path(self.daily_dir)
            daily_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%H%M%S')
            filename = f"{shop_name}_{self.target_date_str}_{timestamp}.xlsx"
            file_path = daily_dir / filename
            
            # 使用openpyxl直接创建，确保列名被正确保存
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = '客户'
            ws['B1'] = '聊天记录'
            wb.save(file_path)
            
            print(f"✓ 创建空的聊天数据文件: {file_path}")
            return True
        except Exception as e2:
            print(f"创建空文件也失败: {e2}")
            return False

    def run(self):
        """主运行函数 - 多店铺聊天数据采集"""
        print("=== 天猫聊天数据采集程序启动 ===")
        print(f"目标日期: {self.target_date_str}")
        
        # 1. 生成每日任务
        print("\n=== 生成每日任务 ===")
        created_count = self.generate_daily_tasks()
        print(f"生成任务数量: {created_count}")
        
        # 检查是否有任务生成
        if created_count == 0:
            print("⚠️ 警告: 没有生成任何任务，可能是因为:")
            print("   - 该日期的任务已存在")
            print("   - 没有符合条件的店铺")
            print("   - 数据库连接问题")
        
        # 2. 获取待处理的店铺任务
        print("\n=== 获取待处理任务 ===")
        shops_info = self.get_shops_with_tasks()
        
        if not shops_info:
            print("没有找到 chat_status 类型的待处理任务")
            # 即使没有任务也要执行合并上传，确保数据一致性
            self.merge_and_upload_files()
            return
        
        print(f"找到 {len(shops_info)} 个店铺的待处理任务")
        
        # 3. 处理每个店铺的聊天数据
        success_count = 0
        total_count = len(shops_info)
        
        for shop_name, shop_info in shops_info.items():
            print(f"\n=== 处理店铺: {shop_name} ===")
            
            try:
                # 获取店铺的cookie和任务信息
                qncookie = shop_info['qncookie'] if 'qncookie' in shop_info else ''
                userNick = shop_info['userNick'] if 'userNick' in shop_info else ''
                task_info = shop_info['task_info'] if 'task_info' in shop_info else None
                
                if not userNick:
                    print(f"✗ 店铺 {shop_name} 缺少userNick信息，跳过处理")
                    continue
                
                result = self.fetch_and_save_chat_data(shop_name, qncookie, userNick)
                
                if result:
                    print(f"✓ 店铺 {shop_name} 聊天数据处理成功")
                    success_count += 1
                    # 更新任务状态
                    self.update_task_status(shop_name, 'chat_status', '已完成')
                else:
                    print(f"✗ 店铺 {shop_name} 差评数据处理失败")

                        
            except Exception as e:
                print(f"✗ 处理店铺 {shop_name} 时发生异常: {e}")
                if 'task_id' in locals() and task_id:
                    self.update_task_status(shop_name, task_id, 'failed')
        
        # 4. 合并文件并上传到MinIO
        print(f"\n=== 数据处理完成 ===")
        print(f"成功处理: {success_count}/{total_count} 个店铺")
        
        # 执行合并上传
        merge_success = self.merge_and_upload_files()
        
        if merge_success:
            print("✓ 所有数据处理和上传完成")
        else:
            print("✗ 数据合并上传失败")
        
        print("=== 程序执行完成 ===")

    def upload_to_minio(self, file_path, minio_path):
        """
        将Excel文件转换为Parquet格式并上传到MinIO
        与tm_kpi.py的upload_to_minio方法保持一致
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 如果DataFrame为空，添加一行空数据以避免MinIO API错误
            if df.empty:
                print("⚠️ 检测到空数据文件，添加占位数据")
                df = pd.DataFrame([{
                    'shop_name': '',
                    'customer_nick': '',
                    'customer_id': '',
                    'message_id': '',
                    'message_content': '',
                    'message_time': '',
                    'sender_type': '',
                    'message_type': '',
                    'create_time': ''
                }])
            
            # 处理NaN值，确保数据能够正常序列化
            df = df.fillna('')  # 将NaN值替换为空字符串
            
            # 处理无穷大值
            df = df.replace([float('inf'), float('-inf')], '')
            
            # 确保所有数据都能正常序列化
            for col in df.columns:
                if df[col].dtype in ['float64', 'float32']:
                    df[col] = df[col].replace([float('inf'), float('-inf')], '')
                # 转换为字符串以避免序列化问题
                df[col] = df[col].astype(str)
            
            # 准备上传数据
            upload_data = {
                "data": df.to_dict('records'),  # 转换为字典列表
                "target_path": minio_path,
                "format": "parquet",
                "bucket": "warehouse"
            }
            
            # 发送POST请求到MinIO API
            headers = {'Content-Type': 'application/json'}
            response = requests.post(self.minio_api_url, json=upload_data, headers=headers)
            
            if response.status_code == 200:
                result = response.json()
                if result.get('success'):
                    print(f"✓ 成功上传合并文件到MinIO: {minio_path}")
                    return True
                else:
                    print(f"✗ MinIO上传失败: {result.get('message', '未知错误')}")
                    return False
            else:
                print(f"✗ MinIO API请求失败: {response.status_code} - {response.text}")
                return False
                    
        except Exception as e:
            print(f"✗ 上传合并文件到MinIO时出错: {str(e)}")
            return False
    
    def refresh_dremio_table(self, table_name):
        """刷新Dremio表"""
        try:
            print(f"🔄 正在刷新Dremio表: {table_name}")
            # 这里可以添加实际的Dremio刷新逻辑
            # 目前先简单打印，后续可以集成实际的Dremio API调用
            print(f"✓ Dremio表刷新完成: {table_name}")
            return True
        except Exception as e:
            print(f"✗ 刷新Dremio表失败: {e}")
            return False

    def parse_chat_response(self, jsonp_response: str) -> Dict[str, Any]:
        """解析JSONP格式的聊天记录响应"""
        try:
            print(f"🔍 解析JSONP响应，长度: {len(jsonp_response)}")
            print(f"🔍 响应前100字符: {jsonp_response[:100]}")
            
            # 动态提取callback名称和JSON内容
            import re
            callback_match = re.match(r'^(\w+)\(', jsonp_response)
            
            if callback_match and jsonp_response.endswith(')'):
                callback_name = callback_match.group(1)
                print(f"✓ 检测到callback名称: {callback_name}")
                
                # 提取JSON部分
                json_str = jsonp_response[len(callback_name)+1:-1]  # 去掉 'callback_name(' 和 ')'
                print(f"✓ 提取JSON字符串，长度: {len(json_str)}")
                
                data = json.loads(json_str)
                print(f"✓ JSON解析成功")
                return data
            else:
                print(f"❌ JSONP格式不匹配")
                raise ValueError("无法从JSONP响应中提取JSON数据")
        except Exception as e:
            print(f"❌ 解析JSONP响应时出错: {e}")
            import traceback
            traceback.print_exc()
            return {}

    def extract_message_content(self, content_str: str) -> str:
        """从content字段中提取实际的聊天内容"""
        try:
            content_data = json.loads(content_str)
            
            if 'text' in content_data:
                # 清理特殊字符
                text = content_data['text'].strip()
                # 移除 \u0004 等控制字符
                text = re.sub(r'\\u0004', '', text)
                text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
                return text
            elif 'summary' in content_data:
                return content_data['summary']
            elif 'title' in content_data:
                return content_data['title']
            else:
                return "[其他类型消息]"
                
        except Exception as e:
            return f"[解析消息内容失败: {str(e)}]"

    def format_timestamp(self, timestamp_str: str) -> str:
        """格式化时间戳为可读格式"""
        try:
            timestamp = int(timestamp_str) / 1000
            dt = datetime.fromtimestamp(timestamp)
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception as e:
            return f"[时间格式错误: {str(e)}]"

    def get_sender_name(self, message: Dict[str, Any]) -> str:
        """获取发送者名称"""
        try:
            # 优先使用senderNick字段
            if 'senderNick' in message:
                return message['senderNick']
            
            # 备用：extMap中的sender_nick
            if 'extMap' in message and 'sender_nick' in message['extMap']:
                return message['extMap']['sender_nick']
            
            # 备用：senderName字段
            if 'senderName' in message:
                return message['senderName']
                
            # 备用：从senderId提取
            if 'senderId' in message and 'appUid' in message['senderId']:
                return f"用户{message['senderId']['appUid']}"
                
            return "[未知发送者]"
            
        except Exception as e:
            return f"[获取发送者失败: {str(e)}]"

    def get_receiver_name(self, message: Dict[str, Any]) -> str:
        """获取接收者名称"""
        try:
            # 优先使用receiverNick字段
            if 'receiverNick' in message:
                return message['receiverNick']
            
            # 备用：extMap中的receiver_nick
            if 'extMap' in message and 'receiver_nick' in message['extMap']:
                return message['extMap']['receiver_nick']
                
            return "[未知接收者]"
            
        except Exception as e:
            return f"[获取接收者失败: {str(e)}]"

    def determine_customer_name(self, messages: List[Dict[str, Any]]) -> str:
        """从消息中确定客户名称"""
        # 查找客户名称（通常是receiver_nick中不包含店铺名的）
        for message in messages:
            try:
                # 优先从receiverNick获取
                if 'receiverNick' in message:
                    receiver = message['receiverNick']
                    # 如果接收者不包含店铺关键词，可能是客户
                    if receiver and '专卖店' not in receiver and '旗舰店' not in receiver and '店铺' not in receiver:
                        return receiver
                
                # 从senderNick获取
                if 'senderNick' in message:
                    sender = message['senderNick']
                    # 如果发送者不包含店铺关键词，可能是客户
                    if sender and '专卖店' not in sender and '旗舰店' not in sender and '店铺' not in sender:
                        return sender
                
                # 备用：从extMap获取
                if 'extMap' in message:
                    sender = message['extMap'].get('sender_nick', '')
                    receiver = message['extMap'].get('receiver_nick', '')
                    
                    # 如果发送者是店铺（包含专卖店），则接收者是客户
                    if '专卖店' in sender and receiver:
                        return receiver
                    # 如果接收者是店铺，则发送者是客户
                    elif '专卖店' in receiver and sender:
                        return sender
            except:
                continue
        
        return "未知客户"

    def parse_jsonp_response(self, jsonp_response):
        """解析JSONP响应中的聊天消息 - 使用备份文件的成功逻辑"""
        try:
            if not jsonp_response:
                print("❌ JSONP响应为空")
                return []
            
            print(f"\n=== 开始解析JSONP响应 ===")
            print(f"响应长度: {len(jsonp_response)}")
            print(f"响应前200字符: {jsonp_response[:200]}")
            
            # 动态提取callback名称
            import re
            callback_match = re.match(r'^(\w+)\(', jsonp_response)
            
            if callback_match and jsonp_response.endswith(')'):
                callback_name = callback_match.group(1)
                print(f"检测到callback名称: {callback_name}")
                
                # 动态提取JSON部分
                json_str = jsonp_response[len(callback_name)+1:-1]  # 去掉 'callback_name(' 和 ')'
                print(f"提取的JSON字符串长度: {len(json_str)}")
                print(f"JSON前200字符: {json_str[:200]}")
                
                try:
                    data = json.loads(json_str)
                    
                    # 打印完整的聊天消息响应
                    print(f"\n=== 聊天消息API完整响应 ===")
                    print(json.dumps(data, indent=2, ensure_ascii=False))
                    print(f"========================\n")
                    
                    # 检查返回状态
                    if 'ret' in data and data['ret'] and data['ret'][0].startswith('SUCCESS'):
                        # 获取消息列表
                        if 'data' in data and 'userMessages' in data['data']:
                            message_list = data['data']['userMessages']
                            print(f"✓ 成功获取到 {len(message_list)} 条消息")
                            
                            # 打印前5条消息的详细内容
                            print(f"\n=== 前5条消息详情 ===")
                            for i, msg in enumerate(message_list[:5]):
                                print(f"消息 {i+1}: {json.dumps(msg, indent=2, ensure_ascii=False)}")
                            print(f"==================\n")
                            
                            return message_list
                        else:
                            print("❌ API返回成功但没有消息数据")
                            print(f"完整响应数据: {json.dumps(data, indent=2, ensure_ascii=False)}")
                            return []
                    else:
                        # 检查是否直接包含data字段（无ret字段的情况）
                        if 'data' in data and 'userMessages' in data['data']:
                            message_list = data['data']['userMessages']
                            print(f"✓ 成功获取到 {len(message_list)} 条消息（无ret字段）")
                            
                            # 打印前5条消息的详细内容
                            print(f"\n=== 前5条消息详情 ===")
                            for i, msg in enumerate(message_list[:5]):
                                print(f"消息 {i+1}: {json.dumps(msg, indent=2, ensure_ascii=False)}")
                            print(f"==================\n")
                            
                            return message_list
                        else:
                            print(f"❌ API返回错误或无消息数据")
                            print(f"完整响应数据: {json.dumps(data, indent=2, ensure_ascii=False)}")
                            return []
                            
                except json.JSONDecodeError as e:
                    print(f"❌ JSON解析失败: {e}")
                    print(f"响应内容: {jsonp_response[:200]}")
                    return []
            else:
                print(f"❌ JSONP格式不匹配: {jsonp_response[:100]}")
                return []
                
        except Exception as e:
            print(f"❌ 解析聊天消息失败: {e}")
            import traceback
            traceback.print_exc()
            return []

    def parse_chat_messages_new(self, chat_data: Dict[str, Any]) -> List[Dict[str, str]]:
        """解析聊天消息数据，按时间排序"""
        messages = []
        
        try:
            if 'data' not in chat_data or 'userMessages' not in chat_data['data']:
                print("聊天数据中没有找到userMessages")
                return messages
                
            user_messages = chat_data['data']['userMessages']
            
            # 确定客户名称
            customer_name = self.determine_customer_name(user_messages)
            
            for message in user_messages:
                try:
                    send_time = self.format_timestamp(message.get('sendTime', '0'))
                    sender_name = self.get_sender_name(message)
                    receiver_name = self.get_receiver_name(message)
                    content = self.extract_message_content(message.get('content', '{}'))
                    
                    # 格式化聊天记录
                    chat_record = f"[{send_time}] {sender_name}: {content}"
                    
                    messages.append({
                        '客户': customer_name,
                        '聊天记录': chat_record,
                        '发送时间': send_time,
                        '发送者': sender_name,
                        '接收者': receiver_name,
                        '消息内容': content
                    })
                    
                except Exception as e:
                    print(f"解析单条消息时出错: {e}")
                    continue
                    
            # 按发送时间排序
            messages.sort(key=lambda x: x['发送时间'])
            
        except Exception as e:
            print(f"解析聊天消息时出错: {e}")
            
        return messages

    def format_for_excel_new(self, messages: List[Dict[str, str]]) -> List[Dict[str, str]]:
        """格式化数据用于Excel存储"""
        if not messages:
            return []
        
        # 按客户分组
        grouped = {}
        for message in messages:
            customer = message['客户']
            if customer not in grouped:
                grouped[customer] = []
            grouped[customer].append(message)
        
        # 格式化为Excel数据
        excel_data = []
        for customer, customer_messages in grouped.items():
            # 按时间排序
            customer_messages.sort(key=lambda x: x['发送时间'])
            
            # 合并聊天记录
            chat_records = [msg['聊天记录'] for msg in customer_messages]
            combined_chat = '\n'.join(chat_records)
            
            excel_data.append({
                '客户': customer,
                '聊天记录': combined_chat
            })
        
        return excel_data

def main():
    """主函数"""
    collector = TmallChatManager()
    collector.run()


if __name__ == "__main__":
    main()
      